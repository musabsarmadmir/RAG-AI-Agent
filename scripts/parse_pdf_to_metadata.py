"""Parse a PDF (sample-files/Fatima_AI_Services-1.pdf) and write a provider metadata.xlsx

Usage:
  python scripts/parse_pdf_to_metadata.py --provider Fatima --pdf sample-files/Fatima_AI_Services-1.pdf

The script extracts service lines from the PDF and writes a one-row Excel file with columns:
  name, email, phone, services_summary, charges

"""
from pathlib import Path
import sys
import argparse
import re
import pandas as pd
from PyPDF2 import PdfReader
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Ensure repo root is on sys.path so `app` package imports work when running this script directly
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.utils import ensure_provider_dirs
from app.config import PROVIDERS_DIR


def extract_text_from_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = []
    for p in reader.pages:
        try:
            pages.append(p.extract_text() or '')
        except Exception:
            # fallback: try to read raw content
            pages.append('')
    return '\n'.join(pages).strip()


def extract_services(text: str):
    # Try to find numbered list items or bullet-like lines
    services = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # match lines like '1. Service name' or '1) Service' or '- Service' or '• Service'
        m = re.match(r'^(?:\d+|[\-\u2022])\s*[\.)\-]?\s*(.+)', line)
        if m:
            services.append(m.group(1).strip())
            continue
        # match numbered '1 ServiceName' or lines starting with number then text
        m2 = re.match(r'^(\d+)\s+([A-Za-z].+)', line)
        if m2:
            services.append(m2.group(2).strip())
            continue
    # If no structured list found, pick lines that look like service headings (longer than 20 chars)
    if not services:
        for line in text.splitlines():
            line = line.strip()
            if len(line) > 20 and not line.endswith(':'):
                services.append(line)
    # Deduplicate while preserving order
    seen = set()
    out = []
    for s in services:
        if s not in seen:
            out.append(s)
            seen.add(s)
    return out


def build_metadata_row(provider: str, services: list):
    # Basic metadata fields used by the app
    name = f"{provider} Test Provider"
    email = f"contact@{provider}.example"
    phone = '555-0100'
    services_summary = '\n'.join(services)
    charges = 'See provider pricing' if services else ''
    return {
        'name': name,
        'email': email,
        'phone': phone,
        'services_summary': services_summary,
        'charges': charges,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--provider', '-p', default='Fatima')
    ap.add_argument('--pdf', default='sample-files/Fatima_AI_Services-1.pdf')
    args = ap.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        print('PDF not found:', pdf_path)
        raise SystemExit(1)

    text = extract_text_from_pdf(pdf_path)
    services = extract_services(text)
    print(f'Extracted {len(services)} service lines')
    for i, s in enumerate(services[:10], 1):
        print(f'  {i}. {s}')

    metadata = build_metadata_row(args.provider, services)
    dirs = ensure_provider_dirs(PROVIDERS_DIR, args.provider)
    out_path = dirs['excel'] / 'metadata.xlsx'
    df = pd.DataFrame([metadata])
    df.to_excel(out_path, index=False)
    print('Wrote metadata to', out_path)

    # Also write into sample-files/Aimdtalk.xlsx under sheet '<provider>_parsed'
    target = Path('sample-files') / 'Aimdtalk.xlsx'
    sheet_name = f"{args.provider}_parsed"

    # Backup existing target file
    if target.exists():
        backups_dir = target.parent / 'backups'
        backups_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        backup_path = backups_dir / f'Aimdtalk.{ts}.xlsx'
        shutil.copy2(str(target), str(backup_path))
        print('Backed up existing Aimdtalk.xlsx to', backup_path)

    # Write dataframe to target sheet, preserving other sheets
    if target.exists():
        # Load workbook and remove existing sheet if present
        wb = load_workbook(str(target))
        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std)
            wb.save(str(target))
        with pd.ExcelWriter(str(target), engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # Create new workbook and write sheet
        with pd.ExcelWriter(str(target), engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print('Wrote parsed metadata to', target, 'sheet', sheet_name)

    # Simple formatting: set column widths and wrap services_summary
    try:
        wb = load_workbook(str(target))
        ws = wb[sheet_name]
        # Adjust columns: name (A), email (B), phone (C), services_summary (D), charges (E)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 25
        # Wrap text for services_summary column (D)
        for cell in ws['D']:
            cell.alignment = Alignment(wrap_text=True)
        wb.save(str(target))
        print('Applied simple formatting to', target, 'sheet', sheet_name)
    except Exception as e:
        print('Warning: failed to apply formatting to', target, e)


if __name__ == '__main__':
    main()
